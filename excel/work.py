#!/usr/bin/env python
# -*- coding:utf-8 -*-
# author:wangtaihe
# datetime:2020/11/16 11:34
# software: PyCharm

import openpyxl as op


def loadWb(path):
    '''
    加载excel文件
    :param path:
    :return:
    '''
    workbook = op.load_workbook(path)
    return workbook


def readSheet(wb, sheet_name):
    '''
    获得sheet
    :param wb:
    :param sheet_name:
    :return:
    '''
    try:
        sheet_by_name = wb[sheet_name]
        return sheet_by_name
    except Exception as e:
        print(e)
        return None


def read_cell(sheet, row, column):
    '''
    读取单元格数据
    :param sheet:
    :param row:
    :param column:
    :return:
    '''
    cell = sheet.cell(row, column)
    return cell.value


def alter_cell(sheet, row, column, out_value):
    sheet.cell(row, column).value = out_value


def read_config_excel(path):
    '''
    根据提供的修改项修改
    :param path:
    :return:
    '''
    sheet = loadWb(path)['Sheet1']
    res = []
    for i in range(1, sheet.max_row - 1):
        sheet_index = sheet.cell(i + 1, 2).value
        row_index = sheet.cell(i + 1, 4).value
        column_index = sheet.cell(i + 1, 5).value
        value = sheet.cell(i + 1, 6).value
        if row_index != None:
            res.append((sheet_index, row_index, column_index, value))
    return res


if __name__ == '__main__':
    configs = read_config_excel('自变量参数.xlsx')
    print(configs)

    wb = loadWb('估值示例(1).xlsx')

    for i in configs:
        sheet = readSheet(wb, i[0])
        if sheet == None:
            print(str(i[0]) + '不存在')
            continue
        alter_cell(sheet, i[1], i[2], i[3])

    wb.save("out.xlsx")
